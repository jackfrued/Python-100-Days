"""
读取Excel文件

Version: 0.1
Author: 骆昊
Date: 2018-03-26
"""

from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('./res/学生明细表.xlsx')
print(workbook.sheetnames)
sheet = workbook[workbook.sheetnames[0]]
print(sheet.title)
for row in range(2, 7):
    for col in range(65, 70):
        cell_index = chr(col) + str(row)
        print(sheet[cell_index].value, end='\t')
    print()
